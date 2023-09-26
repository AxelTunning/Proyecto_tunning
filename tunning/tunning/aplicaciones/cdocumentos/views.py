from django.shortcuts import render,redirect
from django.http import HttpResponseRedirect,HttpResponse
from django.views.generic import (TemplateView, ListView, DetailView, CreateView)
from django.views.generic.edit import FormView
from django.urls import reverse_lazy,reverse
from django.db import transaction
from django.http import JsonResponse
from django.views import View
from django.shortcuts import redirect
from django.urls import reverse
from openpyxl import load_workbook
import logging
from django.contrib import messages

from .forms import ProyectoPersonaForm, DetalleProyectoForm
#**** Nuevo Aqui
import json
#from django.shortcuts import render, redirect
#from .models import Pat_cdoc_header_proyectos, Pat_cdoc_det_proyectos
#from .forms import ProyectoPersonaForm, DetalleProyectoForm

# hasta Aqui

from django.db.models import Q
from .models import *
from .forms import *

# Create your views here.

### Pantalla de Inicio
class Inicio_vista(TemplateView):
    ## Vista que carga pagina de inicio
    template_name = 'incio.html'
    

### Muestra los Proyectos que tiene el Jefe de Proyectos
class Detalle_Proyecto(ListView):
    """ Detalle de Proyecto por Jefe de Proyecto """
    template_name   = 'cdocumentos/JP/list_det_proyectos.html'
    
    def get_queryset(self):
        xcod_jp = self.kwargs['codjp']
        lista   = proyectos_consulta.objects.filter(
                        Id_usuario=xcod_jp)
        return lista
    
    
    
##### Nuevo 

class RegistrarProyectoPersonaView(View):
    template_name = 'cdocumentos/JP/reg_proyectos_asig.html'
    form_proyecto_persona_class = ProyectoPersonaForm
    #success_url = reverse_lazy('cdocumentos_app:proyecto_registrados')
    

    def get(self, request, *args, **kwargs):
        codigo_proyecto = request.GET.get('codigo', 'CódigoDelProyecto')
        form_proyecto_persona = self.form_proyecto_persona_class(initial={'CodigoProyecto': codigo_proyecto})
        nombres_personas = Listado_personal.objects.all()
        return render(request, self.template_name, {
            'form_proyecto_persona': form_proyecto_persona,
            'nombres_personas': nombres_personas,
            'codigo_proyecto': codigo_proyecto,
            'id_jefe_proyecto': 480,
        })

    def post(self, request, *args, **kwargs):
        form_proyecto_persona = self.form_proyecto_persona_class(request.POST)

        if form_proyecto_persona.is_valid():
            proyecto_persona = form_proyecto_persona.save(commit=False)
            proyecto_persona.id_jefe_proyecto = 480  # Asignar el valor aquí
            proyecto_persona.save()

            fichas_seleccionadas_str = request.POST['personalSeleccionado']
            fichas_seleccionadas_list = fichas_seleccionadas_str.strip('[]').split(',')

            for ficha in fichas_seleccionadas_list:
                detalle_proyecto = DetalleProyectoForm({
                    'CodigoProyecto': proyecto_persona.CodigoProyecto,
                    'equipo_proyecto': ficha.strip('\"')
                })
                
                if detalle_proyecto.is_valid():
                    detalle_proyecto.save()
                else:
                    print('********* ERRORES DETALLE PROYECTO *************')
                    print(detalle_proyecto.errors)

            # Obtenemos el valor de 'pk' del proyecto_persona guardado
            pk = proyecto_persona.id_jefe_proyecto
            
            # Generamos la URL inversa usando el nombre de la URL y el 'namespace'
            url = reverse('cdocumentos_app:proyecto_registrados', args=[pk])
            print('Ruta de direccion....:',url)
            
            
            #logging.info(f'Redireccionando a la URL: {url}')  # Agrega esta línea para registrar la redirección
            
            # Redirigir a la URL generada
            return redirect(url)

        else:
            print('********* ERRORES FORMULARIO PROYECTO PERSONA *************')
            print(form_proyecto_persona.errors)

        return render(request, self.template_name, {
            'form_proyecto_persona': form_proyecto_persona,
            'nombres_personas': Listado_personal.objects.all(),
            'id_jefe_proyecto': 480,
        })
        

##### Hasta Aqui

        
##Aqui bueno
### Registra los datos del proyecto, persona de contacto y el personal que va participar en el proyecto
# class RegistrarProyectoPersonaView(View):
#     template_name = 'cdocumentos/JP/reg_proyectos_asig.html'
#     form_proyecto_persona_class = ProyectoPersonaForm
#     success_url = reverse_lazy('cdocumentos_app:proyecto_registrados')

#     def get(self, request, *args, **kwargs):
#         codigo_proyecto = request.GET.get('codigo', 'CódigoDelProyecto')
#         form_proyecto_persona = self.form_proyecto_persona_class(initial={'CodigoProyecto': codigo_proyecto})
#         nombres_personas = Listado_personal.objects.all()
#         return render(request, self.template_name, {
#             'form_proyecto_persona': form_proyecto_persona,
#             'nombres_personas': nombres_personas,
#             'codigo_proyecto': codigo_proyecto,
#             'id_jefe_proyecto': 480,
#         })

#     def post(self, request, *args, **kwargs):
#         form_proyecto_persona = self.form_proyecto_persona_class(request.POST)

#         if form_proyecto_persona.is_valid():
#             proyecto_persona = form_proyecto_persona.save(commit=False)
#             proyecto_persona.id_jefe_proyecto = 480  # Asignar el valor aquí
#             proyecto_persona.save()

#             fichas_seleccionadas_str = request.POST['personalSeleccionado']
#             fichas_seleccionadas_list = fichas_seleccionadas_str.strip('[]').split(',')

#             for ficha in fichas_seleccionadas_list:
#                 detalle_proyecto = DetalleProyectoForm({
#                     'CodigoProyecto': proyecto_persona.CodigoProyecto,
#                     'equipo_proyecto': ficha.strip('\"')
#                 })
                
#                 if detalle_proyecto.is_valid():
#                     detalle_proyecto.save()
#                 else:
#                     print('********* ERRORES DETALLE PROYECTO *************')
#                     print(detalle_proyecto.errors)
#             #return redirect('cdocumentos_app:inicio')
#             # Genera la URL inversa con el argumento codjp
#             codjp = 123  # Reemplaza esto con el valor real de codjp
#             url = reverse('cdocumentos_app:proyecto_asignado', args=[codjp])
#             return redirect(self.success_url)

#         else:
#             print('********* ERRORES FORMULARIO PROYECTO PERSONA *************')
#             print(form_proyecto_persona.errors)

#         return render(request, self.template_name, {
#             'form_proyecto_persona': form_proyecto_persona,
#             'nombres_personas': Listado_personal.objects.all(),
#             'id_jefe_proyecto': 480,
#         })
        
#Hasta Aqui



### Muestra los proyectos que registro el Jefe de Proyecto
class consulta_proyecto_registrado(ListView):
    template_name = 'cdocumentos/JP/list_proyectos_reg.html'
    model = consulta_proyecto_doc
    context_object_name = 'consulta_proyecto'
    
    def get_queryset(self):
         pk = self.kwargs['pk']  # Cambia 'codjp' a 'pk'
         #print("Valor de pk:", pk)  # Agrega esta impresión para verificar el valor de pk
         
         lista = consulta_proyecto_doc.objects.filter(
             id_jefe_proyecto=pk)  # Filtra por 'id_jefe_proyecto'
         
         print("Registros filtrados:", lista)  # Agrega esta impresión para verificar los registros filtrados
         
         
         return lista



# class carga_listado_entregable(View):
#     template_name = 'cdocumentos/JP/carga_excel.html'

#     def post(self, request, *args, **kwargs):
#         excel_file = request.FILES['excelFile']

#         # Procesar el archivo Excel
#         wb = load_workbook(excel_file)
#         ws = wb.active

#         datos = []
#         for row in ws.iter_rows(min_row=2, values_only=True):  # Suponiendo que la primera fila contiene encabezados
#             columna1, columna2 = row[0], row[1]
#             datos.append({'columna1': columna1, 'columna2': columna2})
            
#             # Guardar en la base de datos
#             Pat_cdoc_listado.objects.create(CodigoProyecto=columna1, cod_plano_tunning=columna2)

#         return JsonResponse({'datos': datos})

class ListaEntregables(ListView):
    template_name = 'cdocumentos/JP/carga_excel.html'
    model = Pat_cdoc_listado
    context_object_name = 'entregables'  # Nombre para acceder a la lista de objetos en la plantilla
    
    
    
# Nuevo Aqui seccion de Excel

from django.shortcuts import render
import pandas as pd
from io import BytesIO

def show_excel_table(request):
    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']
        
        # Read Excel data using pandas
        df = pd.read_excel(excel_file)
  
        
        # Convert DataFrame to HTML table
        table_html = df.to_html(index=False, classes='table table-bordered')
        
        return render(request, 'cdocumentos/JP/show_excel_table.html', {'table_html': table_html})
    
    return render(request, 'cdocumentos/JP/carga_excel.html')

